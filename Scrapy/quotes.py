# import scrapy

# class LinkedCompanySpider(scrapy.Spider):
#     name = "linkedin_company_profile"

#     # Add your own list of company URLs here
#     company_pages = [
#         'https://www.linkedin.com/company/usebraintrust?trk=public_jobs_jserp-result_job-search-card-subtitle',
#         'https://www.linkedin.com/company/centraprise?trk=public_jobs_jserp-result_job-search-card-subtitle'
#     ]

#     def start_requests(self):
#         company_index_tracker = 0
#         first_url = self.company_pages[company_index_tracker]
#         yield scrapy.Request(url=first_url, callback=self.parse_response, meta={'company_index_tracker': company_index_tracker})

#     def parse_response(self, response):
#         company_index_tracker = response.meta['company_index_tracker']
#         print('***************')
#         print('****** Scraping page ' + str(company_index_tracker+1) + ' of ' + str(len(self.company_pages)))
#         print('***************')

#         company_item = {}

#         # Extract the basic company info
#         company_item['name'] = response.css('.top-card-layout__entity-info h1::text').get(default='not-found').strip()
#         company_item['summary'] = response.css('.top-card-layout__entity-info h4 span::text').get(default='not-found').strip()

#         try:
#             # Extract the specific company details (industry, size, founded)
#             company_details = response.css('.core-section-container__content .mb-2')

#             # Industry
#             company_industry_line = company_details[1].css('.text-md::text').getall()
#             company_item['industry'] = company_industry_line[1].strip()

#             # Company size
#             company_size_line = company_details[2].css('.text-md::text').getall()
#             company_item['size'] = company_size_line[1].strip()

#             # Founded
#             founded_line = company_details[5].css('.text-md::text').getall()
#             company_item['founded'] = founded_line[1].strip()

#             # Extract additional company details from the <dl> section
#             company_item['website'] = response.css('div[data-test-id="about-us__website"] a::attr(href)').get(default='not-found').strip()
#             company_item['headquarters'] = response.css('div[data-test-id="about-us__headquarters"] dd::text').get(default='not-found').strip()
#             company_item['type'] = response.css('div[data-test-id="about-us__organizationType"] dd::text').get(default='not-found').strip()
#             company_item['specialties'] = response.css('div[data-test-id="about-us__specialties"] dd::text').get(default='not-found').strip()

#             # Extract the location details from divs with id containing "address-"
#             address_divs = response.css('div[id*="address-"]')

#             locations = []
#             for address_div in address_divs:
#                 address_parts = address_div.css('p::text').getall()

#                 # Clean the address parts and check if we have valid data
#                 address_lines = [line.strip() for line in address_parts if line.strip() != '']
                
#                 if address_lines:
#                     locations.append({
#                         'address': address_lines
#                     })

#             # Assign the location to company_item
#             company_item['locations'] = locations

#         except IndexError:
#             print("Error: Skipped Company - Some details missing")

#         yield company_item

#         company_index_tracker = company_index_tracker + 1
#         if company_index_tracker <= (len(self.company_pages) - 1):
#             next_url = self.company_pages[company_index_tracker]
#             yield scrapy.Request(url=next_url, callback=self.parse_response, meta={'company_index_tracker': company_index_tracker})


# import scrapy
# from openpyxl import Workbook
# from openpyxl.utils import get_column_letter

# class LinkedCompanySpider(scrapy.Spider):
#     name = "linkedin_company_profile"

#     company_pages = [
#         'https://www.linkedin.com/company/in-marketing-services/',
#         'https://www.linkedin.com/company/the-marketing-millennials/',
#         'https://www.linkedin.com/company/marketing-week/',
#         'https://www.linkedin.com/company/marketinginspiration/',
#         'https://www.linkedin.com/company/the-marketing-society/',
#         'https://www.linkedin.com/company/the-marketing-meetup/',
#         'https://www.linkedin.com/company/marketing-examples/',
#         'https://www.linkedin.com/company/marketing-przy-kawie/',
#         'https://www.linkedin.com/company/marketingignorante/',
#         'https://www.linkedin.com/company/marketing-weekly/'

#     ]

#     def start_requests(self):
#         company_index_tracker = 0
#         first_url = self.company_pages[company_index_tracker]
#         yield scrapy.Request(url=first_url, callback=self.parse_response, meta={'company_index_tracker': company_index_tracker})

#     def parse_response(self, response):
#         company_index_tracker = response.meta['company_index_tracker']
#         print('***************')
#         print('****** Scraping page ' + str(company_index_tracker+1) + ' of ' + str(len(self.company_pages)))
#         print('***************')

#         company_item = {}

#         # Extract the basic company info
#         company_item['name'] = response.css('.top-card-layout__entity-info h1::text').get(default='not-found').strip()
#         company_item['summary'] = response.css('.top-card-layout__entity-info h4 span::text').get(default='not-found').strip()

#         try:
#             company_details = response.css('.core-section-container__content .mb-2')

#             company_industry_line = company_details[1].css('.text-md::text').getall()
#             company_item['industry'] = company_industry_line[1].strip()

#             company_size_line = company_details[2].css('.text-md::text').getall()
#             company_item['size'] = company_size_line[1].strip()

#             founded_line = company_details[5].css('.text-md::text').getall()
#             company_item['founded'] = founded_line[1].strip()

#             company_item['website'] = response.css('div[data-test-id="about-us__website"] a::attr(href)').get(default='not-found').strip()
#             company_item['headquarters'] = response.css('div[data-test-id="about-us__headquarters"] dd::text').get(default='not-found').strip()
#             company_item['type'] = response.css('div[data-test-id="about-us__organizationType"] dd::text').get(default='not-found').strip()
#             company_item['specialties'] = response.css('div[data-test-id="about-us__specialties"] dd::text').get(default='not-found').strip()

#             # Extract location data and concatenate addresses
#             address_divs = response.css('div[id*="address-"]')

#             locations = []
#             for address_div in address_divs:
#                 address_parts = address_div.css('p::text').getall()

#                 # Strip extra spaces and keep only non-empty address parts
#                 address_lines = [line.strip() for line in address_parts if line.strip() != '']
                
#                 if address_lines:
#                     # Concatenate the address parts into a single string
#                     locations.append(', '.join(address_lines))

#             # Join multiple locations with a comma
#             if locations:
#                 company_item['locations'] = ', '.join(locations)
#             else:
#                 company_item['locations'] = 'not-found'

#         except IndexError:
#             print("Error: Skipped Company - Some details missing")

#         # Save data to Excel (XLSX)
#         self.save_to_excel(company_item)

#         company_index_tracker += 1
#         if company_index_tracker < len(self.company_pages):
#             next_url = self.company_pages[company_index_tracker]
#             yield scrapy.Request(url=next_url, callback=self.parse_response, meta={'company_index_tracker': company_index_tracker})

#     def save_to_excel(self, company_item):
#         # Create a workbook and sheet if not already created
#         try:
#             # If workbook doesn't exist, create a new one
#             self.workbook
#         except AttributeError:
#             self.workbook = Workbook()
#             self.sheet = self.workbook.active
#             self.sheet.title = "Company Profiles"
            
#             # Write headers
#             headers = ['Name', 'Summary', 'Industry', 'Size', 'Founded', 'Website', 'Headquarters', 'Type', 'Specialties', 'Locations']
#             self.sheet.append(headers)

#         # Write company data
#         row = [
#             company_item['Name'],
#             company_item['Summary'],
#             company_item['Industry'],
#             company_item['Size'],
#             company_item['Founded'],
#             company_item['Website'],
#             company_item['Headquarters'],
#             company_item['Type'],
#             company_item['Specialties'],
#             company_item['Locations']  # The locations are now in a single string
#         ]
#         self.sheet.append(row)

#         # Save the workbook
#         self.workbook.save('company_profiles.xlsx')


import scrapy
from openpyxl import load_workbook

class LinkedCompanySpider(scrapy.Spider):
    name = "linkedin_company_profile"

    def __init__(self):
        self.company_pages = self.load_company_urls()

    def load_company_urls(self):
        """Loads company URLs from 'company_data.xlsx' (first 10 rows of 'Company URL' column)."""
        workbook = load_workbook('company_data.xlsx')
        sheet = workbook.active  # Get first sheet
        urls = [sheet.cell(row=i, column=2).value for i in range(2, 21)]  # Read first 10 rows (Column A)
        return [url for url in urls if url]  # Remove empty values

    def start_requests(self):
        for index, url in enumerate(self.company_pages):
            yield scrapy.Request(url=url, callback=self.parse_response, meta={'company_index_tracker': index, 'company_url': url})

    def parse_response(self, response):
        company_index_tracker = response.meta['company_index_tracker']
        company_url = response.meta['company_url']  # Get the company URL from meta
        print(f"***************\n****** Scraping page {company_index_tracker+1} of {len(self.company_pages)}\n***************")

        company_item = {}

        # Extract basic company info
        company_item['Name'] = response.css('.top-card-layout__entity-info h1::text').get(default='Not found').strip()
        company_item['Summary'] = response.css('.top-card-layout__entity-info h4 span::text').get(default='Not found').strip()

        try:
            company_details = response.css('.core-section-container__content .mb-2')

            company_item['Industry'] = company_details[1].css('.text-md::text').getall()[1].strip() if len(company_details) > 1 else 'Not found'
            company_item['Size'] = company_details[2].css('.text-md::text').getall()[1].strip() if len(company_details) > 2 else 'Not found'

            # Handle Founded Year
            if len(company_details) > 5:
                founded_line = company_details[5].css('.text-md::text').getall()
                if len(founded_line) > 1 and founded_line[1].strip().isdigit():
                    company_item['Founded'] = int(founded_line[1].strip())
                else:
                    company_item['Founded'] = 'Not found'
            else:
                company_item['Founded'] = 'Not found'

            company_item['Website'] = response.css('div[data-test-id="about-us__website"] a::attr(href)').get(default='Not found').strip()
            company_item['Headquarters'] = response.css('div[data-test-id="about-us__headquarters"] dd::text').get(default='Not found').strip()
            company_item['Type'] = response.css('div[data-test-id="about-us__organizationType"] dd::text').get(default='Not found').strip()
            company_item['Specialties'] = response.css('div[data-test-id="about-us__specialties"] dd::text').get(default='Not found').strip()

            # Extract location data
            address_divs = response.css('div[id*="address-"]')
            locations = [', '.join([line.strip() for line in address_div.css('p::text').getall() if line.strip()]) for address_div in address_divs]
            company_item['Locations'] = ', '.join(locations) if locations else 'Not found'

            # Extract the "See all employees" URL
            employees_url = response.css('a[data-tracking-control-name="public_biz_employees-join"]::attr(href)').get(default='Not found')
            company_item['Employees'] = employees_url.strip() if employees_url != 'Not found' else 'Not found'

            # Add the company URL to the item
            company_item['Company URL'] = company_url

        except IndexError:
            print("Error: Skipped Company - Some details missing")

        # Save data to Excel
        self.save_to_excel(company_item)

    def save_to_excel(self, company_item):
        try:
            self.workbook
        except AttributeError:
            from openpyxl import Workbook
            self.workbook = Workbook()
            self.sheet = self.workbook.active
            self.sheet.title = "Company Profiles"
            self.sheet.append(['Company URL', 'Name', 'Summary', 'Industry', 'Size', 'Founded', 'Website', 'Headquarters', 'Type', 'Specialties', 'Locations', 'Employees'])

        row = [company_item.get(col, 'Not found') for col in ['Company URL', 'Name', 'Summary', 'Industry', 'Size', 'Founded', 'Website', 'Headquarters', 'Type', 'Specialties', 'Locations', 'Employees']]
        self.sheet.append(row)
        self.workbook.save('company_profilesss.xlsx')
