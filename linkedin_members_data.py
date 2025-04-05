# import requests
# import time
# import csv
# import openpyxl
# from openpyxl import Workbook

# # Google API Key & Custom Search Engine ID
# API_KEY = "AIzaSyBN5V5ZzJwaT3kNRmC9cVZGsMnr3UdU9G4"
# CX = "12ca13bdb666641cd"

# # Function to use Google Custom Search API with pagination by changing the 'start' value in the URL
# def google_search(query, api_key, cx, start_index=0):
#     url = f"https://www.googleapis.com/customsearch/v1?q={query}&key={api_key}&cx={cx}&start={start_index}"
#     response = requests.get(url)
#     return response.json()

# # Function to process the results from the API
# def process_results(data):
#     results = []
#     for item in data.get("items", []):
#         title = item.get("title", "")
#         link = item.get("link", "")
        
#         # Check if the company name is part of the title
#         name = title.replace(" - " + company_name, "").strip() if company_name in title else ""
        
#         if name:
#             results.append((name, link))

#     return results

# # Function to search and process multiple pages with 'start' value changing
# def search_and_print_profiles(query, api_key, cx, max_pages=10):
#     all_results = []
#     start_index = 0  # Initial start index is 0

#     for page in range(max_pages):
#         print(f"Fetching page {page + 1} with start_index = {start_index}...")
        
#         data = google_search(query, api_key, cx, start_index)
        
#         # Check if the 'items' field is available
#         if "items" not in data:
#             print("No more results or API issue.")
#             break
        
#         results = process_results(data)
#         all_results.extend(results)

#         # Increment start_index by 10 for the next page
#         start_index += 10

#         # Sleep to prevent hitting API rate limits
#         time.sleep(2)

#     return all_results

# # Read company names from company_data.xlsx (First 10 rows of the 'Company Name' column)
# company_names = []
# wb = openpyxl.load_workbook("company_data_1.xlsx")
# sheet = wb.active

# for i in range(2, 11):  # Read the first 10 company names
#     company_name = sheet[f"A{i}"].value  # Assuming 'Company Name' is in column A
    # if company_name:
    #     company_names.append(company_name)

# # Prepare the output file for company members
# output_wb = Workbook()
# output_sheet = output_wb.active
# output_sheet.append(["Company Name", "Member Name", "LinkedIn Profile URL"])  # Header row

# # Loop through the company names and scrape data
# for company_name in company_names:
#     print(f"\nScraping data for company: {company_name}")
    
    # # Search query for the current company
    # query = f'"{company_name}" -intitle:"profiles" -inurl:"dir/" site:linkedin.com/in/ OR site:linkedin.com/pub/'
    
#     # Run the search and get profiles for the company
#     profiles = search_and_print_profiles(query, API_KEY, CX)

#     # Write results to the output Excel sheet
#     for name, link in profiles:
#         output_sheet.append([company_name, name, link])

# # Save the results to company_members.xlsx
# output_wb.save("company_membersss.xlsx")

# print(f"\nScraped data has been saved to company_members.xlsx.")



import requests
import time
import openpyxl
import re
from openpyxl import Workbook

# Google API Key & Custom Search Engine ID
API_KEY = "AIzaSyAgGV4pzTPcdrbx5Ne5bb0GW-XYS1B3jok"
CX = "25409c5e0557e48ef"


# Function to use Google Custom Search API with pagination
def google_search(query, api_key, cx, start_index=1):
    url = f"https://www.googleapis.com/customsearch/v1?q={query}&key={api_key}&cx={cx}&start={start_index}"
    response = requests.get(url)
    return response.json()

# Function to process the results from the API
def process_results(data, company_name):
    results = []
    for item in data.get("items", []):
        title = item.get("title", "")
        link = item.get("link", "")

        if "LinkedIn" not in title:
            continue

        # Remove anything after '|'
        title_clean = title.split("|")[0].strip()

        # Skip titles that are just 'LinkedIn', 'Airbnb', etc.
        if title_clean.lower() in ["linkedin", "airbnb", "gerente"]:
            continue

        # Remove company name (case-insensitive)
        title_clean = re.sub(re.escape(company_name), "", title_clean, flags=re.IGNORECASE).strip(" -")

        # Split into name and title
        if " - " in title_clean:
            name, title_part = title_clean.split(" - ", 1)
        else:
            name = title_clean
            title_part = ""

        name = name.strip()
        title_part = title_part.strip()

        if name:
            results.append((name, title_part, link))

    return results

# Function to search and gather multiple pages
def search_and_print_profiles(query, api_key, cx, company_name, max_pages=3):
    all_results = []
    start_index = 1

    for page in range(max_pages):
        print(f"Fetching page {page + 1} with start_index = {start_index}...")

        data = google_search(query, api_key, cx, start_index)

        if "items" not in data:
            print("No more results or API issue.")
            break

        results = process_results(data, company_name)
        all_results.extend(results)

        start_index += 10
        time.sleep(2)

    return all_results

# Read company names from Excel
company_names = []
wb = openpyxl.load_workbook("company_data_1.xlsx")
sheet = wb.active

for i in range(2, 22):  # Read rows 2 to 20 (i.e., first 19 companies)
    company_name = sheet[f"A{i}"].value
    if company_name:
        company_names.append(company_name)

# Prepare output workbook
output_wb = Workbook()
output_sheet = output_wb.active
output_sheet.append(["Company Name", "Member Name", "Title", "LinkedIn Profile URL"])

# Loop through companies and perform search
for company_name in company_names:
    print(f"\nScraping data for company: {company_name}")
    
    query = f'"{company_name}" -intitle:"profiles" -inurl:"dir/" site:linkedin.com/in/ OR site:linkedin.com/pub/'
    profiles = search_and_print_profiles(query, API_KEY, CX, company_name)

    for name, title, link in profiles:
        output_sheet.append([company_name, name, title, link])

# Save results
output_wb.save("company_membersss.xlsx")
print(f"\nScraped data has been saved to company_members.xlsx.")
